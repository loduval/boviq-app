#!/usr/bin/env python3
"""
BOVIQ — Mise à jour automatique des données de marché
Télécharge les Excel DG AGRI (Commission Européenne) et génère cours-data.json
Utilisé par GitHub Actions (hebdomadaire) ou manuellement.
"""

import json, os, sys, urllib.request, tempfile
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# === CONFIG ===
MILK_URL = "https://agriculture.ec.europa.eu/document/download/62d01488-33a0-4601-a841-ca48fa11d999_en?filename=eu-milk-historical-price-series_en.xlsx"
BEEF_URL = "https://agriculture.ec.europa.eu/document/download/db5b282e-a2b4-4d18-91bd-bd4854dc2030_en?filename=bovine-carcase-prices-latest_en.xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(PROJECT_DIR, "data", "market", "cours-data.json")

# Local fallbacks (if download fails)
LOCAL_MILK = os.path.join(PROJECT_DIR, "data", "market", "milk-prices.xlsx")
LOCAL_BEEF = os.path.join(PROJECT_DIR, "data", "market", "beef-prices.xlsx")


def download(url, label):
    """Download file to temp, return path. Returns None on failure."""
    print(f"  📥 Downloading {label}...")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "BOVIQ-MarketData/1.0"})
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        with urllib.request.urlopen(req, timeout=60) as resp:
            tmp.write(resp.read())
        tmp.close()
        size_kb = os.path.getsize(tmp.name) // 1024
        print(f"  ✅ {label}: {size_kb} KB")
        return tmp.name
    except Exception as e:
        print(f"  ⚠️  Download failed ({label}): {e}")
        return None


def extract_milk(path):
    """Extract France milk prices from DG AGRI Excel."""
    print("  🥛 Extracting milk prices...")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Raw Milk Prices"]
    
    # France = col 11, periods in col 1, data from row 8
    # Find France column dynamically
    fr_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(7, c).value
        if v and "france" in str(v).lower():
            fr_col = c
            break
    if not fr_col:
        print("  ⚠️  France column not found in milk sheet")
        wb.close()
        return {"latest": None, "prices": []}
    
    prices = []
    for r in range(8, ws.max_row + 1):
        period = ws.cell(r, 1).value
        price = ws.cell(r, fr_col).value
        if period and price and isinstance(price, (int, float)):
            prices.append({"period": str(period), "price": round(float(price), 2)})
    
    wb.close()
    
    # Keep last 36 months for the app
    prices = prices[-36:] if len(prices) > 36 else prices
    latest = prices[-1] if prices else None
    
    print(f"  ✅ Milk: {len(prices)} months, latest = {latest}")
    return {"latest": latest, "prices": prices}


def extract_beef(path):
    """Extract France beef carcass prices from DG AGRI Excel."""
    print("  🥩 Extracting beef prices...")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Weekly All Carcase Prices"]
    
    # Find France column (row 9 has country codes)
    fr_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(9, c).value
        if v and str(v).strip().upper() == "FR":
            fr_col = c
            break
    if not fr_col:
        print("  ⚠️  France column not found in beef sheet")
        wb.close()
        return {"week_date": "", "prices": {}}
    
    # Extract all categories for France
    category_map = {
        "Young Bulls 12>24m A-U2": "jb_u2",
        "Young Bulls 12>24m A-U3": "jb_u3",
        "Young Bulls 12>24m A-R2": "jb_r2",
        "Young Bulls 12>24m A-R3": "jb_r3",
        "Young Bulls 12>24m A-O2": "jb_o2",
        "Young Bulls 12>24m A-O3": "jb_o3",
        "Bullocks  C-R3": "boeuf_r3",
        "Bullocks  C-O3": "boeuf_o3",
        "Cows D-R3": "vache_r3",
        "Cows D-R4": "vache_r4",
        "Cows D-O2": "vache_o2",
        "Cows D-O3": "vache_o3",
        "Cows D-O4": "vache_o4",
        "Cows D-P2": "vache_p2",
        "Cows D-P3": "vache_p3",
        "Heifers  E-U3": "genisse_u3",
        "Heifers  E-R2": "genisse_r2",
        "Heifers  E-R3": "genisse_r3",
        "Heifers  E-R4": "genisse_r4",
        "Heifers  E-O3": "genisse_o3",
    }
    
    prices = {}
    for r in range(10, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        price = ws.cell(r, fr_col).value
        if cat and price and isinstance(price, (int, float)):
            cat_str = str(cat).strip()
            key = category_map.get(cat_str)
            if key:
                prices[key] = round(float(price), 2)
    
    # Get week date from sheet 1 (scans first rows for dd.mm.yyyy pattern)
    import re
    ws1 = wb["Weekly ACZ Carcase Prices"]
    week_date = ""
    for r in range(1, 7):
        for c in range(1, 40):
            v = ws1.cell(r, c).value
            if v:
                m = re.search(r"(\d{2}\.\d{2}\.\d{4})", str(v))
                if m:
                    week_date = m.group(1)
                    break
        if week_date:
            break
    
    wb.close()
    print(f"  ✅ Beef: {len(prices)} categories, week = {week_date}")
    return {"week_date": week_date, "prices": prices}


def main():
    print("🐄 BOVIQ — Mise à jour des cours du marché")
    print(f"   {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    print()
    
    # Try download, fallback to local files
    milk_path = download(MILK_URL, "Lait") or LOCAL_MILK
    beef_path = download(BEEF_URL, "Viande") or LOCAL_BEEF
    
    if not os.path.exists(milk_path):
        print(f"❌ Milk file not found: {milk_path}")
        sys.exit(1)
    if not os.path.exists(beef_path):
        print(f"❌ Beef file not found: {beef_path}")
        sys.exit(1)
    
    milk = extract_milk(milk_path)
    beef = extract_beef(beef_path)
    
    # Build output JSON
    result = {
        "meta": {
            "generated": datetime.now(timezone.utc).isoformat(),
            "source": "DG AGRI — Commission Européenne",
            "country": "France",
            "milk_unit": "EUR/100kg",
            "beef_unit": "EUR/100kg carcasse"
        },
        "milk": milk,
        "beef": beef
    }
    
    # Write JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"\n✅ cours-data.json généré ({size_kb:.1f} KB)")
    print(f"   → {OUTPUT_PATH}")
    
    # Cleanup temp files
    for p in [milk_path, beef_path]:
        if p.startswith(tempfile.gettempdir()):
            os.unlink(p)


if __name__ == "__main__":
    main()
